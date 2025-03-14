import math
import openpyxl as xls
import numpy as np


def F(x, y):
    '''
        Функция, используемая в программе.
    '''
    return x ** 2 + y ** 2


def parse_table(name):
    '''
        Загрузка таблицы в программу.
    '''
    try:
        pos = 1
        points = xls.load_workbook(name).active
        table = []
        while points.cell(row = pos, column = 1).value is not None:
            buf = []
            for i in range(1, 8):
                buf.append(float(points.cell(row = pos, column = i).value))
            table.append(buf)
            pos += 1 
        print(table)
        # Разбиение на нужные массивы (X, Y, Z)
        x = []
        y = []
        z = []
        for i in range(1, len(table)):
            buf = []
            x.append(table[i][0])
            y.append(table[0][i])
            for j in range(1, len(table)):
                buf.append(table[i][j])
            z.append(buf)
        print(x, y, z)
        return z, x, y
        
    except TypeError:
        print("Проверьте данные на вводе!!!")
        return None, None, None
    except ValueError:
        print("Проверьте данные на вводе!!!")
        return None, None, None


# def find_x0_xn1(data, power, arg):
#     '''
#         Нахождение начального и конечного индекса в таблице (x0 и xn).
#     '''
#     #print("find_x0_xn")
#     index_x = 0
#     len_z = len(data)
    
#     while arg > data[index_x]:
#         index_x += 1
#         if arg < data[index_x]:
#             index_x -= 1
#             break
    
#     index_x0 = index_x - math.ceil(power / 2) + 1
#     index_xn = index_x + math.ceil(power / 2) + power % 2 - 1
#     #print("\n\n", index_x0, index_xn)
#     if (index_x + math.ceil(power / 2) >= len_z):
#         index_xn = len_z
#         index_x0 = len_z - power
#     elif (index_x < math.ceil(power / 2)):
#         index_x0 = 0
#         index_xn = power
#     else:
#         index_x0 = index_x - math.ceil(power / 2) + 1
#         index_xn = index_x0 + power
#     #print("\n\n", index_x0, index_xn)
#     return index_x0, index_xn


def find_x0_xn(data, power, arg):
    '''
        Нахождение начального и конечного индекса в таблице (x/y0 и x/yn).
    '''
   # print("find_x0_xn1")
    index_x = 0
    
    while arg > data[index_x]:
        index_x += 1
        if arg < data[index_x]:
            index_x -= 1
            break
    index_x0 = index_x - math.ceil(power / 2) + 1
    index_xn = index_x + math.ceil(power / 2) + ((power - 1) % 2)
    if index_xn > len(data) - 1:
        index_x0 -= index_xn - len(data) + 1
        index_xn = len(data) - power
    elif index_x0 < 0:
        index_xn += -index_x0
        index_x0 = 0
   # print("\n\n", index_x0, index_xn)
    return index_x0, index_xn


def div_diff(z, node):
    '''
        Расчет разделенных разниц для полинома Ньютона
    '''
    for i in range(node):
        pol = []
        for j in range(node - i):
            buf = (z[i + 1][j] - z[i + 1][j + 1]) / (z[0][j] - z[0][j + i + 1])
            pol.append(buf)
        z.append(pol)
    # for i in z:
    #     print(i)
    return z


def polinom_n(z, node, arg):
    '''
        Расчет значение функции от заданного аргумента.
        Полином Ньютона.
    '''
    pol = div_diff(z, node)
    y = 0
    buf = 1
    for i in range(node + 1):
        y += buf * pol[i + 1][0]
        buf *= (arg - pol[0][i])
    return y 


def multid_interp(z, x, y, power_x, power_y, arg_x, arg_y):
    index_x0, index_xn = find_x0_xn(x, power_x + 1, arg_x)
    index_y0, index_yn = find_x0_xn(y, power_y + 1, arg_y)

    x = x[index_x0:index_xn]
    y = y[index_y0:index_yn]
    z = z[index_y0:index_yn]
    for i in range(power_y + 1):
        z[i] = z[i][index_x0:index_xn]
    #for i in range(power_y + 1):
        #print("qqqqqqqqqqqqq", [x, z[i]])
    x1 = [polinom_n([x, z[i]], power_x, arg_x) for i in range(power_y + 1)]
    print("x1",x1)
    y1 = polinom_n([y, x1], power_y, arg_y)
    return y1


def input_xy():
    '''
        Ввод аргумента. (в случае ошибки дается еще попытка)
    '''
    print("Enter X: ")
    flag = 0
    x = 0
    while flag == 0:
        x = input(float)
        try:
            val = float(x)
            flag = 1
        except ValueError:
            print("Some error! Try again")
    print("Enter Y: ")
    flag = 0
    y = 0
    while flag == 0:
        y = input(float)
        try:
            val = float(y)
            flag = 1
        except ValueError:
            print("Some error! Try again")
    return float(x), float(y)

def printf_matrix(z):
    print("\n\n y/x", end="   0   1   2   3   4   5\n")
    for i in range(len(z)):
        print("%4d" % i, end="")
        for j in range(len(z)):
            print("%4d" % z[i][j], end="")
        print()

def main():
    z, x, y = parse_table("c:/msys64/home/zolot/VA/lab2/points.xlsx")
    arg_x, arg_y = input_xy()
    arr_n = [1, 2, 3]
    printf_matrix(z)

    print("\n|  nx  |  ny  |   x   |   y   | Found Y | F(x, y) | Error  |")
    for n in arr_n:
        found_y = multid_interp(z, x, y, n, n, arg_x, arg_y)
        
        print("|   %d  |   %d  |  %.2f |  %.2f | %.5f | %.5f | %.4f |" \
            % (n, n, arg_x, arg_y, found_y, F(arg_x, arg_y), 
            abs(found_y - F(arg_x, arg_y))))


if __name__ == "__main__":
    main()

